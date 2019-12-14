import openpyxl
import smtplib

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.active
lastcol = sheet.max_column
latestmonth = sheet.cell(row = 1, column = lastcol).value
unpaidmembers = {}
for r in range(2,sheet.max_row+1):
    payment = sheet.cell(row = r, column = lastcol).value
    if payment != 'paid':
        name = sheet.cell(row = r, column = 1).value
        email = sheet.cell(row = r, column = 2).value
        unpaidmembers[name] = email
print(unpaidmembers)
smtpobj = smtplib.SMTP('smtp.gmail.com', 587)
smtpobj.ehlo()
smtpobj.starttls()
smtpobj.login(#enter your email,#enter you password)
#for name,email in unpaidmembers.items():
 #   body = 'Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid for %s. Please make this payment as soon as possible. Thankyou' %(latestmonth, name, latestmonth)
  #  print('Sending out mail to %s...' %name)
   # sendmailstatus = smtpobj.sendmail(#enter your email, email, body)
    #if sendmailstatus != {}:
     #   print('There was a problem sending mail to %s: %s' %(name,sendmailstatus))
smtpobj.quit()
