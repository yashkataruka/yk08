import openpyxl
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
price_updates={'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}
for row in range(2, sheet.max_row+1):
    productname = sheet.cell(row=row,column = 1).value
    if productname in price_updates:
        sheet.cell(row=row,column = 2).value = price_updates[productname]
wb.save('updated_produceSales.xlsx')
