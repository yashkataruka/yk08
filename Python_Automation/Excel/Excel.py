import os
#os.chdir('C:\\users\\yashk\\Desktop')
import openpyxl
#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb['Sheet1']
#print(sheet.max_row)
#print(sheet.max_column)
#from openpyxl.utils import get_column_letter,column_index_from_string
#print(get_column_letter(1))
#print(get_column_letter(900))
#print(column_index_from_string('SSS'))
#print(get_column_letter(sheet.max_column))
#print(tuple(sheet['A1':'C3']))
#print(sheet['A1':'C3'])
#for rowOfCellObjects in sheet['A1':'C3']:
#    for cellObj in rowOfCellObjects:
 #       print(cellObj.coordinate)
  #      print(cellObj.value)
   # print('--- END OF ROW ---')
#print(sheet['A1':'C3'])
#sheet = wb.active chooses the sheet which opens default
#sheet = wb['Sheet1']
#print(sheet)
#print(sheet['B'])
#for i in sheet['B']:
 #   print(i.value)


#wb = openpyxl.Workbook()
#print(wb.sheetnames)
#sheet = wb.active
#print(sheet.title)
#sheet.title = 'Spam Bacon Eggs Sheet'
#print(wb.sheetnames)


#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb['Sheet1']
#sheet.title = 'Spam Spam Spam'
#wb.save('example_copy.xlsx')


#wb = openpyxl.Workbook()
#wb.create_sheet()
#print(wb.sheetnames)
#wb.create_sheet(index=0,title='First Sheet')
#print(wb.sheetnames)
#wb.create_sheet(index = 2, title = 'Middle Sheet')
#print(wb.sheetnames)
#wb.remove(wb['First Sheet'])
#print(wb.sheetnames)


#sheet = wb['Sheet']
#sheet['A1'] = 'Hello World!'
#print(sheet['A1'].value)
#wb.save('A_random.xlsx')


#from openpyxl.styles import Font
#wb = openpyxl.Workbook()
#sheet = wb['Sheet']
#italic24font = Font(size = 24, italic = True)
#sheet['A1'].font = italic24font
#sheet['A1'] = 'Hello World'
#fontObj = Font(name = 'Times New Roman', bold=True)
#sheet['B1'].font = fontObj
#sheet['B1'] = 'Bold Times New Roman'
#fontObj1 = Font(size = 24, italic = True, bold=True)
#sheet['B3'].font = fontObj1
#sheet['B3'] = '24 pt. italic and bold'
#wb.save('styled.xlsx')


#sheet['A1'] = 200
#sheet['A2'] = 300
#sheet['A3'] = '=SUM(A1:A2)'
#wb.save('writeFormula.xlsx')


#sheet['A1'] = 'Tall row'
#sheet['B2'] = 'Wide column'
#sheet.row_dimensions[1].height = 70
#sheet.column_dimensions['B'].width = 20
#wb.save('dimensions.xlsx')


#sheet.merge_cells('A1:D3')
#sheet['A1'] = 'Twelve cells merged together.'
#sheet.merge_cells('C5:D5')
#sheet['C5'] = 'Two merged cells'
#sheet['E5'] = 'Two merged cells'
#sheet.merge_cells('E5:F5')
#wb.save('merged.xlsx')
#wb = openpyxl.load_workbook('merged.xlsx')
#sheet = wb.active
#sheet.unmerge_cells('A1:D3')
#sheet.unmerge_cells('C5:D5')
#sheet.unmerge_cells('E5:F5')
#wb.save('merged.xlsx')


#wb = openpyxl.load_workbook('produceSales.xlsx')
#sheet = wb.active
#sheet.freeze_panes = 'A3'
#sheet.freeze_panes = 'C3'
#wb.save('freezeExample.xlsx')


wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1,11):
    sheet['A'+str(i)]=i
refobj = openpyxl.chart.Reference(sheet, min_col = 1, min_row=1, max_col=1, max_row=20)
seriesobj = openpyxl.chart.Series(refobj, title = 'First series')
chartobj = openpyxl.chart.BarChart()
chartobj.title = 'My chart'
chartobj.append(seriesobj)
sheet.add_chart(chartobj,'C5')
wb.save('sampleChart.xlsx')
